from flask import Flask, render_template, request, send_file, abort
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import openpyxl
from openpyxl.styles import numbers
import io

app = Flask(__name__)

def setup_driver():
    service = Service()
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    return webdriver.Chrome(service=service, options=options)

def dismiss_popup(driver):
    try:
        dismiss_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='dismiss-button']"))
        )
        dismiss_button.click()
        print("Popup dismissed")
    except TimeoutException:
        print("No popup found or dismiss button not clickable")

def login_to_ubersuggest(driver, email, password):
    try:
        driver.get("https://app.neilpatel.com/en/login")
        wait = WebDriverWait(driver, 20)
        
        # Dismiss initial popup if it appears
        dismiss_popup(driver)
        
        email_input = wait.until(EC.presence_of_element_located((By.NAME, "email")))
        email_input.send_keys(email)
        
        password_input = driver.find_element(By.NAME, "password")
        password_input.send_keys(password)
        
        login_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='login-button']")))
        login_button.click()
        
        # Wait for login to complete and redirect to dashboard
        wait.until(EC.url_contains("dashboard"))
        
        # Dismiss popup after login if it appears
        dismiss_popup(driver)
        
        print("Login successful")
    except Exception as e:
        print(f"Login failed: {str(e)}")
        raise

def analyze_keyword(driver, keyword):
    try:
        url = f"https://app.neilpatel.com/en/ubersuggest/keyword_ideas?ai-keyword={keyword}&keyword={keyword}&lang=en&locId=2826&mode=keyword"
        driver.get(url)
        
        wait = WebDriverWait(driver, 20)
        
        volume_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "sc-dUWDJJ.bywnwD")))
        search_volume = volume_element.text.replace(',', '')
        
        seo_difficulty_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "sc-eiQriw.ewoFEM")))
        seo_difficulty = seo_difficulty_element.text

        print(f"Analyzed {keyword}: Volume = {search_volume}, SEO Difficulty = {seo_difficulty}")
        return {"Keyword": keyword, "Search Volume": search_volume, "SEO Difficulty": seo_difficulty}

    except TimeoutException:
        print(f"Timeout occurred while analyzing keyword: {keyword}")
        return {"Keyword": keyword, "Search Volume": "N/A", "SEO Difficulty": "N/A"}
    except Exception as e:
        print(f"Error occurred while analyzing keyword: {keyword}. Error: {str(e)}")
        return {"Keyword": keyword, "Search Volume": "Error", "SEO Difficulty": "Error"}

def create_xlsx(results):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Keyword Analysis"

    ws['A1'] = "Keyword"
    ws['B1'] = "Volume"
    ws['C1'] = "SEO Difficulty"

    for row, result in enumerate(results, start=2):
        ws[f'A{row}'] = result['Keyword']
        
        try:
            ws[f'B{row}'] = int(result['Search Volume'])
            ws[f'B{row}'].number_format = numbers.FORMAT_NUMBER
        except ValueError:
            ws[f'B{row}'] = result['Search Volume']
        
        try:
            ws[f'C{row}'] = int(result['SEO Difficulty'])
            ws[f'C{row}'].number_format = numbers.FORMAT_NUMBER
        except ValueError:
            ws[f'C{row}'] = result['SEO Difficulty']

    wb.save(output)
    output.seek(0)
    return output

def process_keywords(input_file, email, password):
    df = pd.read_excel(input_file)
    keywords = df['Keyword'].tolist()

    driver = setup_driver()
    
    try:
        login_to_ubersuggest(driver, email, password)
        results = []

        for keyword in keywords:
            print(f"Analyzing keyword: {keyword}")
            result = analyze_keyword(driver, keyword)
            results.append(result)
            time.sleep(5)

        return create_xlsx(results)
    finally:
        driver.quit()

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    try:
        if request.method == 'POST':
            if 'file' not in request.files:
                return 'No file part'
            file = request.files['file']
            email = request.form['email']
            password = request.form['password']
            if file.filename == '':
                return 'No selected file'
            if file and file.filename.endswith('.xlsx'):
                output = process_keywords(file, email, password)
                return send_file(
                    output,
                    as_attachment=True,
                    download_name='keyword_analysis_results.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
        return render_template('upload.html')
    except Exception as e:
        app.logger.error(f"An error occurred: {str(e)}")
        return f"An error occurred: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True, port=8000)
